"""
Report exporters for LakChogo Connect
Supports CSV, PDF, and Excel formats
"""

import csv
import json
from datetime import datetime
from django.http import HttpResponse
import io

# Try to import optional libraries
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportExporter:
    """Handle exporting reports to different formats"""
    
    def __init__(self, data, title='Report', format_type='csv'):
        self.data = data
        self.title = title
        self.format_type = format_type
    
    def export(self):
        """Export report based on format type"""
        if self.format_type == 'pdf':
            return self.to_pdf()
        elif self.format_type == 'excel':
            return self.to_excel()
        else:
            return self.to_csv()
    
    def to_csv(self):
        """Export to CSV format"""
        response = HttpResponse(content_type='text/csv')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_title = self.title.replace(' ', '_').replace('/', '_')
        response['Content-Disposition'] = f'attachment; filename="{safe_title}_{timestamp}.csv"'
        
        writer = csv.writer(response)
        self._write_data(writer)
        return response
    
    def to_pdf(self):
        """Export to PDF format"""
        if not REPORTLAB_AVAILABLE:
            return self._fallback_response('PDF export requires reportlab library. Install with: pip install reportlab')
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_title = self.title.replace(' ', '_').replace('/', '_')
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{safe_title}_{timestamp}.pdf"'
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph(self.title, title_style))
        elements.append(Spacer(1, 20))
        
        # Get headers and rows
        headers, rows = self._get_data()
        
        if headers and rows:
            # Prepare table data
            table_data = [headers]
            table_data.extend(rows)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table)
            
            # Add summary
            if 'summary' in self.data:
                elements.append(Spacer(1, 20))
                summary_style = ParagraphStyle(
                    'Summary',
                    parent=styles['Heading2'],
                    fontSize=14,
                    spaceAfter=10
                )
                elements.append(Paragraph('Summary', summary_style))
                
                for key, value in self.data['summary'].items():
                    elements.append(Paragraph(f"{key.replace('_', ' ').title()}: {value}", styles['Normal']))
        
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response
    
    def to_excel(self):
        """Export to Excel format"""
        if not OPENPYXL_AVAILABLE:
            return self._fallback_response('Excel export requires openpyxl library. Install with: pip install openpyxl')
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_title = self.title.replace(' ', '_').replace('/', '_')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{safe_title}_{timestamp}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = self.title[:31]  # Excel sheet name max 31 chars
        
        # Get headers and rows
        headers, rows = self._get_data()
        
        if headers:
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        if 'summary' in self.data:
            summary_ws = wb.create_sheet('Summary')
            row = 1
            for key, value in self.data['summary'].items():
                summary_ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                summary_ws.cell(row=row, column=2, value=value)
                row += 1
        
        wb.save(response)
        return response
    
    def _get_data(self):
        """Extract headers and rows from data"""
        headers = []
        rows = []
        
        if isinstance(self.data, dict):
            if 'headers' in self.data and 'rows' in self.data:
                headers = self.data['headers']
                rows = self.data['rows']
            else:
                # Convert dict to table
                headers = ['Metric', 'Value']
                for key, value in self.data.items():
                    if not isinstance(value, (list, dict)):
                        rows.append([key.replace('_', ' ').title(), str(value)])
        
        return headers, rows
    
    def _write_data(self, writer):
        """Write data to CSV writer"""
        headers, rows = self._get_data()
        
        if headers:
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            
            # Add summary
            if 'summary' in self.data:
                writer.writerow([])
                writer.writerow(['SUMMARY'])
                for key, value in self.data['summary'].items():
                    writer.writerow([key.replace('_', ' ').title(), value])
        else:
            writer.writerow(['Message', 'No data available'])
    
    def _fallback_response(self, message):
        """Return a simple text response when required library is not available"""
        response = HttpResponse(content_type='text/plain')
        response.write(f"Error: {message}")
        return response
